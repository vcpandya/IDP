"""Parsers for bulk-send recipient lists: CSV, XLSX, and pasted tables.

All parsers return a tuple ``(headers, rows)`` where ``rows`` is a list of dicts
keyed by the original header strings. Empty leading/trailing whitespace is
trimmed; entirely-blank rows are dropped.
"""

from __future__ import annotations

import csv
import io
from typing import List, Tuple

MAX_ROWS = 5000  # bulk-send hard cap; protects email provider + DB


def _clean_headers(raw: List[str]) -> List[str]:
    out = []
    seen: dict[str, int] = {}
    for h in raw:
        h_clean = (h or "").strip()
        if not h_clean:
            h_clean = f"col_{len(out) + 1}"
        # disambiguate duplicates
        if h_clean in seen:
            seen[h_clean] += 1
            h_clean = f"{h_clean}_{seen[h_clean]}"
        else:
            seen[h_clean] = 1
        out.append(h_clean)
    return out


def _rows_from_iter(headers: List[str], data_iter) -> List[dict]:
    rows: List[dict] = []
    for raw in data_iter:
        if raw is None:
            continue
        cells = [("" if c is None else str(c)).strip() for c in raw]
        if not any(cells):
            continue
        # pad short rows / clip long rows
        if len(cells) < len(headers):
            cells = cells + [""] * (len(headers) - len(cells))
        elif len(cells) > len(headers):
            cells = cells[: len(headers)]
        rows.append({headers[i]: cells[i] for i in range(len(headers))})
        if len(rows) >= MAX_ROWS:
            break
    return rows


def parse_csv(data: bytes | str) -> Tuple[List[str], List[dict]]:
    """Parse a CSV byte string or text into (headers, rows)."""
    if isinstance(data, bytes):
        # try utf-8 then latin-1 fallback
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("latin-1", errors="replace")
    else:
        text = data
    reader = csv.reader(io.StringIO(text))
    try:
        first = next(reader)
    except StopIteration:
        return [], []
    headers = _clean_headers(first)
    rows = _rows_from_iter(headers, reader)
    return headers, rows


def parse_xlsx(data: bytes) -> Tuple[List[str], List[dict]]:
    """Parse the first sheet of an XLSX workbook into (headers, rows)."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError("openpyxl is required to parse .xlsx files") from e
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        return [], []
    headers = _clean_headers([("" if c is None else str(c)) for c in first])
    rows = _rows_from_iter(headers, rows_iter)
    return headers, rows


def parse_paste(text: str) -> Tuple[List[str], List[dict]]:
    """Parse pasted spreadsheet content (TSV from Excel/Sheets, or CSV)."""
    if not text or not text.strip():
        return [], []
    # Detect delimiter on the header line
    first_line = text.splitlines()[0] if text else ""
    if "\t" in first_line:
        delim = "\t"
    elif "," in first_line:
        delim = ","
    elif ";" in first_line:
        delim = ";"
    else:
        delim = "\t"
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    try:
        first = next(reader)
    except StopIteration:
        return [], []
    headers = _clean_headers(first)
    rows = _rows_from_iter(headers, reader)
    return headers, rows


def parse_recipients(filename: str | None, content: bytes | None, paste_text: str | None) -> Tuple[List[str], List[dict], str]:
    """Dispatch to the right parser. Returns (headers, rows, source_label)."""
    if content and filename:
        name = filename.lower()
        if name.endswith(".xlsx"):
            h, r = parse_xlsx(content)
            return h, r, filename
        if name.endswith(".csv") or name.endswith(".txt"):
            h, r = parse_csv(content)
            return h, r, filename
        # try CSV as a last resort
        h, r = parse_csv(content)
        return h, r, filename
    if paste_text:
        h, r = parse_paste(paste_text)
        return h, r, f"Pasted ({len(r)} rows)"
    return [], [], ""

"""Spreadsheet parser for XLSX and CSV files."""

import csv
import logging
import os

from .base import BaseParser, ParseResult

logger = logging.getLogger(__name__)


class SpreadsheetParser(BaseParser):
    """Extract text from spreadsheet files (XLSX and CSV).

    Uses ``openpyxl`` for ``.xlsx`` files and the built-in :mod:`csv` module
    for ``.csv`` files.  Each sheet (or the single CSV) becomes a separate
    "page" in the :class:`ParseResult`.
    """

    def supported_extensions(self) -> list[str]:
        return ["xlsx", "xls", "csv"]

    def parse(self, file_path: str) -> ParseResult:
        """Parse a spreadsheet file and return its text content.

        Parameters
        ----------
        file_path:
            Path to the ``.xlsx``, ``.xls``, or ``.csv`` file.

        Raises
        ------
        ImportError
            If openpyxl is not installed and the file is an Excel format.
        """
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            return self._parse_csv(file_path)
        else:
            return self._parse_xlsx(file_path)

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def _parse_csv(self, file_path: str) -> ParseResult:
        logger.info("Parsing CSV: %s", file_path)

        rows: list[list[str]] = []
        with open(file_path, "r", encoding="utf-8", errors="replace", newline="") as fh:
            reader = csv.reader(fh)
            for row in reader:
                rows.append(row)

        text = _rows_to_text(rows, sheet_name="Sheet1")
        pages = [{"page": 1, "text": text}]

        metadata = {
            "format": "csv",
            "row_count": len(rows),
            "column_count": len(rows[0]) if rows else 0,
        }

        logger.info("CSV parsing complete: %d rows", len(rows))

        return ParseResult(
            text=text,
            pages=pages,
            metadata=metadata,
            page_count=1,
        )

    # ------------------------------------------------------------------
    # XLSX / XLS
    # ------------------------------------------------------------------

    def _parse_xlsx(self, file_path: str) -> ParseResult:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX parsing. "
                "Install it with: pip install openpyxl"
            )

        logger.info("Parsing XLSX: %s", file_path)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        pages: list[dict] = []
        all_text_parts: list[str] = []
        sheet_names = wb.sheetnames

        for page_num, sheet_name in enumerate(sheet_names, start=1):
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                str_row = [str(cell) if cell is not None else "" for cell in row]
                rows.append(str_row)

            sheet_text = _rows_to_text(rows, sheet_name=sheet_name)
            pages.append({"page": page_num, "text": sheet_text})
            all_text_parts.append(sheet_text)

        wb.close()

        full_text = "\n\n".join(all_text_parts)
        page_count = len(pages)

        metadata = {
            "format": "xlsx",
            "sheet_count": page_count,
            "sheet_names": sheet_names,
        }

        logger.info("XLSX parsing complete: %d sheets extracted", page_count)

        return ParseResult(
            text=full_text,
            pages=pages,
            metadata=metadata,
            page_count=page_count,
        )


def _rows_to_text(rows: list[list[str]], sheet_name: str = "") -> str:
    """Convert a list of rows into readable text with a markdown table."""
    if not rows:
        return f"## {sheet_name}\n\n(empty sheet)" if sheet_name else "(empty)"

    lines: list[str] = []
    if sheet_name:
        lines.append(f"## {sheet_name}")
        lines.append("")

    # Build markdown table
    max_cols = max(len(r) for r in rows)
    header = rows[0] + [""] * (max_cols - len(rows[0]))
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")
    for row in rows[1:]:
        padded = row + [""] * (max_cols - len(row))
        lines.append("| " + " | ".join(padded[:max_cols]) + " |")

    return "\n".join(lines)
